# -*- coding: utf-8 -*-
"""
Created on Sun Nov 13 19:58:18 2022

@author: lucia
"""
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook



path = 'agenda.xlsx'

wb = openpyxl.load_workbook(path)

cont = 0
hoja = wb.active
for i in range(0, hoja.max_row): #leer datos iniciales del excel
    for columna in hoja.iter_cols(1, hoja.max_column):
        print(columna[i].value, end="\t\t")
        cont+=1
    print('')

print("Ahora introduzca los datos de la nueva persona:  ")
nombre = input("Intro nombre:")
apellido = input("Intro apellido: ")
numero = input("Intro numero: ")
n =5
var1 = hoja.cell(row=n, column=1)
var1.value = nombre
var2 = hoja.cell(row=n, column=2)
var2.value = apellido
var3 = hoja.cell(row=n, column=3)
var3.value = numero

print("Listo!  Compruebe su excel.")
wb.save('agenda.xlsx')