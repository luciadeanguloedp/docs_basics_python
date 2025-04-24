# -*- coding: utf-8 -*-
"""
Created on Sun Nov 13 18:08:52 2022
AGENDA - NOMBRE,APELLIDO,TELF
@author: lucia
"""


import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
cont = 0

wookbook = openpyxl.load_workbook("agenda.xlsx")


worksheet = wookbook.active
book = Workbook()
sheet = book.active

sheet['A1'] = 'Nombre'
sheet['B1'] = 'Apellido'
sheet['C1'] = 'Número'
sheet['A1'].font= Font(color='FF0000', bold=True)
sheet['B1'].font= Font(color='FF0000', bold=True)
sheet['C1'].font= Font(color='FF0000', bold=True)

for i in range(0, worksheet.max_row): #leer datos iniciales del excel
    for columna in worksheet.iter_cols(1, worksheet.max_column):
        print(columna[i].value, end="\t\t")
        cont+=1
    print('')
#ahora quiero añadir una fila con otra persona
print("Hay", int(cont/3-1), "personas")
print("Introduzca los datos de la nueva persona: ")
nom = input("Nombre: ")
ap = input("pellido: ")
num = input("Introduzca un número: ")

sheet['A5']= nom
sheet['B4']= ap
sheet['C5']= num

book.save('ejemplo.xlsx')